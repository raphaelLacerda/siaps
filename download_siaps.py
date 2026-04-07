#!/usr/bin/env python3
"""
Script para baixar dados do SIAPS e gerar arquivos CSV e XLSX
Lê configurações do arquivo equipes_indicadores.json
"""

import requests
import json
import sys
import re
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Aviso: openpyxl não instalado. XLSX não será gerado.")

# Diretório base
BASE_DIR = Path(__file__).parent
DOWNLOADS_DIR = BASE_DIR / "downloads"
CSV_DIR = DOWNLOADS_DIR / "csv"
XLSX_DIR = DOWNLOADS_DIR / "xlsx"


def ensure_directories():
    """Cria diretórios de download se não existirem"""
    CSV_DIR.mkdir(parents=True, exist_ok=True)
    XLSX_DIR.mkdir(parents=True, exist_ok=True)


def get_token():
    """Obtém token do .env"""
    env_path = BASE_DIR / ".env"
    with open(env_path, "r") as f:
        for line in f:
            line = line.strip()
            if line.startswith("bearer_token="):
                return line.split("=", 1)[1]
    raise ValueError("Token não encontrado no .env")


def load_config():
    """Carrega configurações do arquivo equipes_indicadores.json"""
    config_path = BASE_DIR / "equipes_indicadores.json"
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def slugify(text: str) -> str:
    """Converte texto para slug (lowercase, hífens)
    Ex: 'Pessoa Idosa' -> 'pessoa-idosa'
    Ex: 'IST (HIV/Sífilis/Hepatites B e C)' -> 'ist-hiv-sifilis-hepatites-b-e-c'
    """
    # Remove acentos
    acentos = {
        "á": "a",
        "à": "a",
        "ã": "a",
        "â": "a",
        "ä": "a",
        "é": "e",
        "è": "e",
        "ê": "e",
        "ë": "e",
        "í": "i",
        "ì": "i",
        "î": "i",
        "ï": "i",
        "ó": "o",
        "ò": "o",
        "õ": "o",
        "ô": "o",
        "ö": "o",
        "ú": "u",
        "ù": "u",
        "û": "u",
        "ü": "u",
        "ç": "c",
        "ñ": "n",
        "Á": "a",
        "À": "a",
        "Ã": "a",
        "Â": "a",
        "Ä": "a",
        "É": "e",
        "È": "e",
        "Ê": "e",
        "Ë": "e",
        "Í": "i",
        "Ì": "i",
        "Î": "i",
        "Ï": "i",
        "Ó": "o",
        "Ò": "o",
        "Õ": "o",
        "Ô": "o",
        "Ö": "o",
        "Ú": "u",
        "Ù": "u",
        "Û": "u",
        "Ü": "u",
        "Ç": "c",
        "Ñ": "n",
    }
    for char, replacement in acentos.items():
        text = text.replace(char, replacement)

    # Lowercase
    text = text.lower()

    # Substitui caracteres especiais por espaço
    text = re.sub(r"[^a-z0-9\s]", " ", text)

    # Substitui múltiplos espaços por um
    text = re.sub(r"\s+", " ", text).strip()

    # Substitui espaços por hífen
    text = text.replace(" ", "-")

    return text


def format_equipes_name(equipes: list) -> str:
    """Formata lista de equipes para nome do arquivo"""
    return "-".join(equipes)


def format_competencia(competencia: str) -> str:
    """Converte '2025-04' para '202504'"""
    return competencia.replace("-", "")


def format_competencia_display(competencia: str) -> str:
    """Converte 202504 para ABR/25"""
    meses = {
        "01": "JAN",
        "02": "FEV",
        "03": "MAR",
        "04": "ABR",
        "05": "MAI",
        "06": "JUN",
        "07": "JUL",
        "08": "AGO",
        "09": "SET",
        "10": "OUT",
        "11": "NOV",
        "12": "DEZ",
    }
    comp = competencia.replace("-", "")
    ano = comp[2:4]
    mes = comp[4:6]
    return f"{meses[mes]}/{ano}"


def get_total_records(
    competencia: str, indicador: int, equipes: list, token: str
) -> int:
    """Faz request inicial para saber o total de registros"""
    base_url = "https://apisiaps.saude.gov.br/componente/qualidade/visao-competencia"

    params = {
        "page": 0,
        "size": 5,
        "search": "",
        "sort": "string",
        "competencias": format_competencia(competencia),
        "coMunicipioIbge": "530010",
        "indicadores": indicador,
        "sgEquipes": equipes,
    }

    headers = {"Authorization": token}

    response = requests.get(base_url, params=params, headers=headers)
    response.raise_for_status()

    data = response.json()
    return data.get("total", 0)


def fetch_data(
    competencia: str, indicador: int, equipes: list, total: int, token: str
) -> list:
    """Busca todos os dados"""
    base_url = "https://apisiaps.saude.gov.br/componente/qualidade/visao-competencia"

    params = {
        "page": 0,
        "size": total + 100,  # Margem de segurança
        "search": "",
        "sort": "string",
        "competencias": format_competencia(competencia),
        "coMunicipioIbge": "530010",
        "indicadores": indicador,
        "sgEquipes": equipes,
    }

    headers = {"Authorization": token}

    response = requests.get(base_url, params=params, headers=headers)
    response.raise_for_status()

    data = response.json()
    return data.get("content", [])


def process_data(raw_data: list) -> list:
    """Processa os dados JSON para formato tabular"""
    rows = []
    for item in raw_data:
        for equipe in item.get("equipes", []):
            for indicador in equipe.get("indicadores", []):
                row = {
                    "CNES": equipe.get("coCnes", ""),
                    "ESTABELECIMENTO": equipe.get("noUnidade", ""),
                    "TIPO DO ESTABELECIMENTO": equipe.get("dsTipoUnidade", ""),
                    "INE": equipe.get("coEquipe", ""),
                    "NOME DA EQUIPE": equipe.get("noEquipe", ""),
                    "SIGLA DA EQUIPE": equipe.get("sgEquipe", ""),
                    "NÚMERO TOTAL DE ATENDIMENTOS POR DEMANDA PROGRAMADA": indicador.get(
                        "numerador", 0
                    ),
                    "NÚMERO TOTAL DE ATENDIMENTOS POR TODOS OS TIPOS DE DEMANDAS (ESPONTÂNEAS E PROGRAMADAS)": indicador.get(
                        "denominador", 0
                    ),
                    "PONTUAÇÃO": indicador.get("scoreFormatado", "0,00"),
                }
                rows.append(row)
    return rows


def generate_header(competencia: str, indicador_nome: str, equipes: list) -> list:
    """Gera o cabeçalho do relatório"""
    now = datetime.now()

    meses_pt = {
        1: "janeiro",
        2: "fevereiro",
        3: "março",
        4: "abril",
        5: "maio",
        6: "junho",
        7: "julho",
        8: "agosto",
        9: "setembro",
        10: "outubro",
        11: "novembro",
        12: "dezembro",
    }

    data_geracao = f"{now.day:02d} de {meses_pt[now.month]} de {now.year} - {now.hour:02d}:{now.minute:02d}h"
    comp_display = format_competencia_display(competencia)
    equipes_display = ", ".join(equipes)

    return [
        "Ministério da Saúde MS",
        "Secretaria de Atenção Primária à Saúde SAPS",
        "Sistema de Informação para a Atenção Primária à Saúde – SIAPS",
        f"Dado gerado em: {data_geracao}",
        "Relatório Qualidade - Visão por Competência",
        "Dado Preliminar",
        "",
        "Dados Sócio Demográficos:",
        "UF: DF",
        "Município: 530010 / BRASÍLIA",
        "",
        "Filtro:",
        f"Indicador: {indicador_nome}",
        f"Competência selecionada: {comp_display}",
        "Condição das Equipes: Todas as equipes do Município",
        f"Tipo de Equipe: {equipes_display}",
        "",
    ]


def write_csv(
    rows: list, competencia: str, indicador_nome: str, equipes: list, output_path: Path
):
    """Escreve arquivo CSV"""
    header_lines = generate_header(competencia, indicador_nome, equipes)

    columns = [
        "CNES",
        "ESTABELECIMENTO",
        "TIPO DO ESTABELECIMENTO",
        "INE",
        "NOME DA EQUIPE",
        "SIGLA DA EQUIPE",
        "NÚMERO TOTAL DE ATENDIMENTOS POR DEMANDA PROGRAMADA",
        "NÚMERO TOTAL DE ATENDIMENTOS POR TODOS OS TIPOS DE DEMANDAS (ESPONTÂNEAS E PROGRAMADAS)",
        "PONTUAÇÃO",
    ]

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        for line in header_lines:
            f.write(line + "\n")

        f.write(";".join(columns) + "\n")

        for row in rows:
            values = []
            for col in columns:
                val = str(row.get(col, ""))
                values.append(f'"{val}\t"')
            f.write(";".join(values) + "\n")

    print(f"    CSV: {output_path.name}")


def write_xlsx(
    rows: list, competencia: str, indicador_nome: str, equipes: list, output_path: Path
):
    """Escreve arquivo XLSX"""
    if not HAS_OPENPYXL:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"

    header_lines = generate_header(competencia, indicador_nome, equipes)

    for i, line in enumerate(header_lines, 1):
        ws.cell(row=i, column=1, value=line)

    columns = [
        "CNES",
        "ESTABELECIMENTO",
        "TIPO DO ESTABELECIMENTO",
        "INE",
        "NOME DA EQUIPE",
        "SIGLA DA EQUIPE",
        "NÚMERO TOTAL DE ATENDIMENTOS POR DEMANDA PROGRAMADA",
        "NÚMERO TOTAL DE ATENDIMENTOS POR TODOS OS TIPOS DE DEMANDAS (ESPONTÂNEAS E PROGRAMADAS)",
        "PONTUAÇÃO",
    ]

    header_row = len(header_lines) + 1
    for j, col in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=j, value=col)
        cell.font = Font(bold=True)

    for i, row in enumerate(rows, header_row + 1):
        for j, col in enumerate(columns, 1):
            val = row.get(col, "")
            ws.cell(row=i, column=j, value=val)

    for col_idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(
            15, len(col_name) // 2
        )

    wb.save(output_path)
    print(f"    XLSX: {output_path.name}")


def download_single(
    competencia: str, indicador: dict, equipes: list, token: str
) -> bool:
    """Baixa dados de uma competência, indicador e equipes específicos"""
    indicador_nome = indicador["nome"]
    indicador_codigo = indicador["codigo"]
    indicador_slug = slugify(indicador_nome)
    equipes_name = format_equipes_name(equipes)

    # Formatar competência para nome do arquivo (2025-04)
    comp_format = (
        competencia if "-" in competencia else f"{competencia[:4]}-{competencia[4:]}"
    )

    # Nome do arquivo: <equipe>-<indicador>-<competencia>-relatorio-competencia
    base_name = f"{equipes_name}-{indicador_slug}-{comp_format}-relatorio-competencia"

    print(
        f"\n  [{indicador_nome}] Equipes: {equipes_name} | Competência: {comp_format}"
    )

    try:
        # Primeiro, verificar total de registros
        print(f"    Verificando total de registros...")
        total = get_total_records(competencia, indicador_codigo, equipes, token)

        if total == 0:
            print(f"    Nenhum dado encontrado")
            return False

        print(f"    Total encontrado: {total} registros")

        # Buscar todos os dados
        raw_data = fetch_data(competencia, indicador_codigo, equipes, total, token)
        rows = process_data(raw_data)
        print(f"    Processados: {len(rows)} linhas")

        # Salvar arquivos
        write_csv(
            rows, competencia, indicador_nome, equipes, CSV_DIR / f"{base_name}.csv"
        )
        write_xlsx(
            rows, competencia, indicador_nome, equipes, XLSX_DIR / f"{base_name}.xlsx"
        )

        return True
    except requests.exceptions.HTTPError as e:
        print(f"    Erro HTTP: {e.response.status_code} - {e.response.text[:100]}")
        return False
    except Exception as e:
        print(f"    Erro: {e}")
        return False


def download_all(equipes_filter: list = None):
    """Baixa todos os indicadores e competências do arquivo de configuração.
    Se equipes_filter for informado, baixa apenas os grupos que contenham
    ao menos uma das equipes listadas.
    """
    ensure_directories()
    token = get_token()
    config = load_config()

    competencias = config.get("competencias", [])
    equipes_config = config.get("equipes", [])

    if equipes_filter:
        equipes_filter_upper = [e.upper() for e in equipes_filter]
        equipes_config = [
            grupo
            for grupo in equipes_config
            if any(
                sg.upper() in equipes_filter_upper for sg in grupo.get("sgEquipes", [])
            )
        ]
        if not equipes_config:
            print(
                f"Nenhum grupo encontrado para as equipes: {', '.join(equipes_filter)}"
            )
            return

    print("=" * 70)
    print("Download em lote - SIAPS")
    print("=" * 70)
    print(f"Competências: {', '.join(competencias)}")
    print(f"Grupos de equipes: {len(equipes_config)}")
    if equipes_filter:
        print(f"Filtro de equipes: {', '.join(equipes_filter)}")
    print("=" * 70)

    total_downloads = 0
    sucesso = 0

    for grupo in equipes_config:
        equipes = grupo.get("sgEquipes", [])
        indicadores = grupo.get("indicadores", [])

        print(f"\n{'='*70}")
        print(f"EQUIPES: {', '.join(equipes)}")
        print(f"Indicadores: {len(indicadores)}")
        print(f"{'='*70}")

        for indicador in indicadores:
            for competencia in competencias:
                total_downloads += 1
                if download_single(competencia, indicador, equipes, token):
                    sucesso += 1

    print("\n" + "=" * 70)
    print(f"Concluído: {sucesso}/{total_downloads} downloads com sucesso")
    print(f"Arquivos salvos em:")
    print(f"  CSV:  {CSV_DIR}")
    print(f"  XLSX: {XLSX_DIR}")
    print("=" * 70)


def main():
    equipes_filter = sys.argv[1:] if len(sys.argv) > 1 else None
    download_all(equipes_filter)


if __name__ == "__main__":
    main()
